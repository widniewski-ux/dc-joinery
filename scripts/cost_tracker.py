#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import os
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "ops" / "costs"
WORKBOOK_PATH = OUTPUT_DIR / "dcjoineryni-cost-tracker.xlsx"


COST_HEADERS = [
    "Date",
    "Vendor",
    "Service",
    "Category",
    "BillingCycle",
    "UnitCostGBP",
    "Qty",
    "TotalCostGBP",
    "MonthlyEquivalentGBP",
    "Notes",
    "VerificationStatus",
    "Evidence",
]

CATEGORIES = ["Domain", "Hosting", "Email", "AI/API", "Tools", "Other"]
AUX_HEADERS = ["Provider", "Status", "Details", "CheckedAt"]
AUX_AUDIT_PATH = OUTPUT_DIR / "provider-audit.json"


def style_header(ws: Worksheet, row: int, columns: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def month_start(year: int, month: int) -> date:
    return date(year, month, 1)


def next_month_start(year: int, month: int) -> date:
    if month == 12:
        return date(year + 1, 1, 1)
    return date(year, month + 1, 1)


def days_ago(days: int) -> date:
    return date.fromordinal(date.today().toordinal() - days)


def initialize_workbook(
    path: Path,
    initial_rows: Iterable[dict[str, str | float]] | None = None,
    audit_rows: Iterable[dict[str, str]] | None = None,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    costs_ws = wb.active
    costs_ws.title = "Costs"

    for col, header in enumerate(COST_HEADERS, start=1):
        costs_ws.cell(row=1, column=col, value=header)
    style_header(costs_ws, 1, len(COST_HEADERS))

    costs_ws.column_dimensions["A"].width = 13
    costs_ws.column_dimensions["B"].width = 22
    costs_ws.column_dimensions["C"].width = 34
    costs_ws.column_dimensions["D"].width = 14
    costs_ws.column_dimensions["E"].width = 12
    costs_ws.column_dimensions["F"].width = 12
    costs_ws.column_dimensions["G"].width = 8
    costs_ws.column_dimensions["H"].width = 14
    costs_ws.column_dimensions["I"].width = 20
    costs_ws.column_dimensions["J"].width = 42
    costs_ws.column_dimensions["K"].width = 18
    costs_ws.column_dimensions["L"].width = 48

    for row in initial_rows or []:
        append_cost_row(
            costs_ws,
            entry_date=parse_date(str(row["date"])),
            vendor=str(row["vendor"]),
            service=str(row["service"]),
            category=str(row["category"]),
            cycle=str(row["billing_cycle"]),
            unit_cost=float(row["unit_cost_gbp"]),
            qty=float(row.get("qty", 1)),
            notes=str(row.get("notes", "")),
            verification_status=str(row.get("verification_status", "pending")),
            evidence=str(row.get("evidence", "")),
        )

    monthly_ws = wb.create_sheet("Monthly Report")
    monthly_headers = ["Month", "TotalSpentGBP", "Domain", "Hosting", "Email", "AI/API", "Tools", "Other"]
    for col, header in enumerate(monthly_headers, start=1):
        monthly_ws.cell(row=1, column=col, value=header)
    style_header(monthly_ws, 1, len(monthly_headers))

    year = date.today().year
    for idx, month in enumerate(range(1, 13), start=2):
        start = month_start(year, month)
        end = next_month_start(year, month)
        monthly_ws.cell(row=idx, column=1, value=start.strftime("%Y-%m"))
        monthly_ws.cell(
            row=idx,
            column=2,
            value=f'=SUMIFS(Costs!$H:$H,Costs!$A:$A,">={start.isoformat()}",Costs!$A:$A,"<{end.isoformat()}")',
        )
        for offset, category in enumerate(CATEGORIES, start=3):
            monthly_ws.cell(
                row=idx,
                column=offset,
                value=(
                    f'=SUMIFS(Costs!$H:$H,Costs!$A:$A,">={start.isoformat()}",'
                    f'Costs!$A:$A,"<{end.isoformat()}",Costs!$D:$D,"{category}")'
                ),
            )

    monthly_ws.column_dimensions["A"].width = 12
    for col in "BCDEFGH":
        monthly_ws.column_dimensions[col].width = 15

    rolling_ws = wb.create_sheet("Last 4 Weeks")
    rolling_headers = ["WindowStart", "WindowEnd", "TotalSpentGBP"]
    for col, header in enumerate(rolling_headers, start=1):
        rolling_ws.cell(row=1, column=col, value=header)
    style_header(rolling_ws, 1, len(rolling_headers))

    for idx in range(4):
        start = days_ago(28 - (idx * 7))
        end = days_ago(21 - (idx * 7))
        row = idx + 2
        rolling_ws.cell(row=row, column=1, value=start.isoformat())
        rolling_ws.cell(row=row, column=2, value=end.isoformat())
        rolling_ws.cell(
            row=row,
            column=3,
            value=f'=SUMIFS(Costs!$H:$H,Costs!$A:$A,">="&A{row},Costs!$A:$A,"<"&B{row})',
        )

    rolling_ws.column_dimensions["A"].width = 14
    rolling_ws.column_dimensions["B"].width = 14
    rolling_ws.column_dimensions["C"].width = 18

    dashboard_ws = wb.create_sheet("Dashboard")
    dashboard_ws["A1"] = "DC Joinery - Cost Dashboard"
    dashboard_ws["A1"].font = Font(size=16, bold=True)
    dashboard_ws["A3"] = "Total project spend"
    dashboard_ws["B3"] = "=SUM(Costs!H:H)"
    dashboard_ws["A4"] = "Monthly recurring equivalent"
    dashboard_ws["B4"] = "=SUM(Costs!I:I)"
    dashboard_ws["A5"] = "Average monthly spend (this year)"
    dashboard_ws["B5"] = "=AVERAGE('Monthly Report'!B2:B13)"
    dashboard_ws["A6"] = "Rolling 4 weeks spend"
    dashboard_ws["B6"] = "=SUM('Last 4 Weeks'!C2:C5)"
    dashboard_ws["A7"] = "Cost by category"
    dashboard_ws["A8"] = "Category"
    dashboard_ws["B8"] = "TotalGBP"
    style_header(dashboard_ws, 8, 2)

    for row, category in enumerate(CATEGORIES, start=9):
        dashboard_ws.cell(row=row, column=1, value=category)
        dashboard_ws.cell(row=row, column=2, value=f'=SUMIFS(Costs!$H:$H,Costs!$D:$D,"{category}")')

    dashboard_ws.column_dimensions["A"].width = 30
    dashboard_ws.column_dimensions["B"].width = 20

    pie = PieChart()
    pie.title = "Cost split by category"
    pie_data = Reference(dashboard_ws, min_col=2, min_row=8, max_row=8 + len(CATEGORIES))
    pie_labels = Reference(dashboard_ws, min_col=1, min_row=9, max_row=8 + len(CATEGORIES))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    pie.height = 7
    pie.width = 10
    dashboard_ws.add_chart(pie, "D3")

    line = LineChart()
    line.title = "Monthly spend trend"
    line.y_axis.title = "GBP"
    line.x_axis.title = "Month"
    line_data = Reference(monthly_ws, min_col=2, min_row=1, max_row=13)
    line_labels = Reference(monthly_ws, min_col=1, min_row=2, max_row=13)
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(line_labels)
    line.height = 7
    line.width = 11
    dashboard_ws.add_chart(line, "D20")

    audit_ws = wb.create_sheet("Audit")
    for col, header in enumerate(AUX_HEADERS, start=1):
        audit_ws.cell(row=1, column=col, value=header)
    style_header(audit_ws, 1, len(AUX_HEADERS))
    audit_ws.column_dimensions["A"].width = 20
    audit_ws.column_dimensions["B"].width = 10
    audit_ws.column_dimensions["C"].width = 100
    audit_ws.column_dimensions["D"].width = 22
    for row in audit_rows or []:
        next_row = audit_ws.max_row + 1
        audit_ws.cell(row=next_row, column=1, value=row.get("provider", ""))
        audit_ws.cell(row=next_row, column=2, value=row.get("status", ""))
        audit_ws.cell(row=next_row, column=3, value=row.get("details", ""))
        audit_ws.cell(row=next_row, column=4, value=row.get("checked_at", ""))

    wb.save(path)


def append_cost_row(
    ws: Worksheet,
    entry_date: date,
    vendor: str,
    service: str,
    category: str,
    cycle: str,
    unit_cost: float,
    qty: float,
    notes: str,
    verification_status: str = "manual",
    evidence: str = "",
) -> None:
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=entry_date.isoformat())
    ws.cell(row=next_row, column=2, value=vendor)
    ws.cell(row=next_row, column=3, value=service)
    ws.cell(row=next_row, column=4, value=category)
    ws.cell(row=next_row, column=5, value=cycle)
    ws.cell(row=next_row, column=6, value=unit_cost)
    ws.cell(row=next_row, column=7, value=qty)
    ws.cell(row=next_row, column=8, value=f"=F{next_row}*G{next_row}")
    ws.cell(
        row=next_row,
        column=9,
        value=(
            f'=IF(E{next_row}="monthly",H{next_row},'
            f'IF(E{next_row}="yearly",H{next_row}/12,0))'
        ),
    )
    ws.cell(row=next_row, column=10, value=notes)
    ws.cell(row=next_row, column=11, value=verification_status)
    ws.cell(row=next_row, column=12, value=evidence)


def ensure_workbook(path: Path) -> None:
    if path.exists():
        return
    initialize_workbook(path)


def parse_date(input_value: str | None) -> date:
    if not input_value:
        return date.today()
    return datetime.strptime(input_value, "%Y-%m-%d").date()


def cmd_init() -> None:
    initial_rows = []
    audit_rows = []
    if AUX_AUDIT_PATH.exists():
        payload = json.loads(AUX_AUDIT_PATH.read_text())
        initial_rows = payload.get("cost_rows", [])
        audit_rows = payload.get("audit_rows", [])
    initialize_workbook(WORKBOOK_PATH, initial_rows=initial_rows, audit_rows=audit_rows)
    print(f"Workbook created: {WORKBOOK_PATH}")


def _http_json(
    url: str, *, method: str = "GET", headers: dict[str, str] | None = None, body: dict | None = None
) -> tuple[int | None, dict]:
    req = urllib.request.Request(url, method=method)
    effective_headers = {"User-Agent": "dc-joinery-cost-sync/1.0", **(headers or {})}
    for key, value in effective_headers.items():
        req.add_header(key, value)
    payload = None
    if body is not None:
        payload = json.dumps(body).encode("utf-8")
        req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, data=payload, timeout=30) as response:
            raw = response.read().decode("utf-8", "ignore")
            parsed = json.loads(raw) if raw else {}
            return response.status, parsed
    except urllib.error.HTTPError as err:
        raw = err.read().decode("utf-8", "ignore")
        try:
            parsed = json.loads(raw) if raw else {}
        except json.JSONDecodeError:
            parsed = {"raw": raw[:1000]}
        return err.code, parsed
    except Exception as err:
        return None, {"error": str(err)}


def _parse_iso_datetime(value: str) -> datetime | None:
    if not value:
        return None
    normalized = value.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


def _probe_resend_usage(api_key: str, period_start: datetime) -> tuple[str, str]:
    status, payload = _http_json("https://api.resend.com/emails", headers={"Authorization": f"Bearer {api_key}"})
    if status != 200:
        return "pending", f"status={status} details={payload}"

    emails = payload.get("data", []) if isinstance(payload, dict) else []
    count_28d = 0
    for item in emails:
        created_at = _parse_iso_datetime(str(item.get("created_at", "")))
        if created_at and created_at >= period_start:
            count_28d += 1
    return "ok", f"emails_28d={count_28d} api_status=200"


def _probe_replicate_usage(api_key: str, period_start: datetime) -> tuple[str, str]:
    status, payload = _http_json(
        "https://api.replicate.com/v1/predictions?limit=100",
        headers={"Authorization": f"Token {api_key}"},
    )
    if status != 200:
        return "pending", f"status={status} details={payload}"

    predictions = payload.get("results", []) if isinstance(payload, dict) else []
    count_28d = 0
    total_predict_seconds = 0.0
    for item in predictions:
        created_at = _parse_iso_datetime(str(item.get("created_at", "")))
        if not created_at or created_at < period_start:
            continue
        count_28d += 1
        metrics = item.get("metrics", {}) or {}
        try:
            total_predict_seconds += float(metrics.get("predict_time", 0) or 0)
        except (TypeError, ValueError):
            pass
    return "ok", f"predictions_28d={count_28d} predict_time_sec_28d={round(total_predict_seconds, 2)}"


def _probe_supabase_usage(supabase_url: str, service_role_key: str, period_start: datetime) -> tuple[str, str]:
    period_start_iso = period_start.isoformat()
    headers = {"apikey": service_role_key, "Authorization": f"Bearer {service_role_key}"}
    jobs_status, jobs_payload = _http_json(
        (
            f"{supabase_url}/rest/v1/ai_design_jobs"
            f"?select=id,created_at&created_at=gte.{urllib.parse.quote(period_start_iso)}"
        ),
        headers=headers,
    )
    if jobs_status != 200:
        return "pending", f"jobs_status={jobs_status} details={jobs_payload}"

    objects_status, objects_payload = _http_json(
        (
            f"{supabase_url}/rest/v1/objects"
            "?select=metadata,created_at,bucket_id&bucket_id=eq.ai-designer"
            f"&created_at=gte.{urllib.parse.quote(period_start_iso)}"
        ),
        headers={**headers, "Accept-Profile": "storage"},
    )
    jobs_count = len(jobs_payload) if isinstance(jobs_payload, list) else 0
    if objects_status != 200:
        return (
            "ok",
            f"jobs_28d={jobs_count} storage_probe_status={objects_status} storage_probe_details={objects_payload}",
        )

    object_count = len(objects_payload) if isinstance(objects_payload, list) else 0
    bytes_total = 0
    for row in objects_payload if isinstance(objects_payload, list) else []:
        metadata = row.get("metadata") if isinstance(row, dict) else None
        if isinstance(metadata, dict):
            try:
                bytes_total += int(metadata.get("size", 0) or 0)
            except (TypeError, ValueError):
                pass
    mb_total = round(bytes_total / (1024 * 1024), 4)
    return "ok", f"jobs_28d={jobs_count} storage_objects_28d={object_count} storage_mb_28d={mb_total}"


def _pdfshift_audit(api_key: str) -> tuple[int | None, dict, str]:
    body = {"source": "<html><body>audit</body></html>"}
    status, payload = _http_json(
        "https://api.pdfshift.io/v3/convert/pdf",
        method="POST",
        headers={"X-API-Key": api_key},
        body=body,
    )
    if status == 401:
        auth = base64.b64encode(f"{api_key}:".encode()).decode()
        status, payload = _http_json(
            "https://api.pdfshift.io/v3/convert/pdf",
            method="POST",
            headers={"Authorization": f"Basic {auth}"},
            body=body,
        )
        return status, payload, "basic-fallback"
    return status, payload, "x-api-key"


def _merge_verified_costs(
    previous_payload: dict,
    audit_rows: list[tuple[str, str, str, str]],
    cost_rows: list[dict[str, str | float]],
) -> tuple[list[tuple[str, str, str, str]], list[dict[str, str | float]]]:
    previous_cost_rows = {
        str(row.get("vendor", "")): row
        for row in previous_payload.get("cost_rows", [])
        if isinstance(row, dict)
    }
    previous_audit_rows = {
        str(row.get("provider", "")): row
        for row in previous_payload.get("audit_rows", [])
        if isinstance(row, dict)
    }

    merged_cost_rows: list[dict[str, str | float]] = []
    for current in cost_rows:
        vendor = str(current.get("vendor", ""))
        previous = previous_cost_rows.get(vendor)
        if (
            previous
            and str(previous.get("verification_status", "")) == "verified"
            and float(previous.get("unit_cost_gbp", 0) or 0) > 0
            and str(current.get("verification_status", "")) != "verified"
        ):
            preserved = dict(previous)
            preserved["notes"] = (
                f"{previous.get('notes', '')} | preserved: latest API audit could not verify automatically."
            ).strip(" |")
            merged_cost_rows.append(preserved)
            continue
        merged_cost_rows.append(current)

    merged_audit_rows: list[tuple[str, str, str, str]] = []
    for provider, status, details, checked_at in audit_rows:
        previous = previous_audit_rows.get(provider)
        if previous and previous.get("status") == "ok" and status in {"pending", "missing", "error"}:
            previous_details = str(previous.get("details", ""))
            latest_segment = f"latest probe: {details}"
            if latest_segment in previous_details:
                merged_details = previous_details
            else:
                merged_details = f"{previous_details} | {latest_segment}"
            merged_audit_rows.append(
                (
                    provider,
                    "ok",
                    merged_details,
                    checked_at,
                )
            )
            continue
        merged_audit_rows.append((provider, status, details, checked_at))

    return merged_audit_rows, merged_cost_rows


def cmd_audit(args: argparse.Namespace) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    env = os.environ.copy()
    env_local = ROOT / ".env.local"
    if env_local.exists():
        for line in env_local.read_text().splitlines():
            if "=" not in line or line.lstrip().startswith("#"):
                continue
            key, value = line.split("=", 1)
            env.setdefault(key.strip(), value.strip().strip('"').strip("'"))

    checked_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    period_days = max(1, int(args.days))
    period_start = datetime.now(timezone.utc) - timedelta(days=period_days)
    previous_payload: dict = {}
    if AUX_AUDIT_PATH.exists():
        try:
            previous_payload = json.loads(AUX_AUDIT_PATH.read_text())
        except json.JSONDecodeError:
            previous_payload = {}
    vercel_token = args.vercel_token or env.get("VERCEL_TOKEN")
    audit_rows: list[tuple[str, str, str, str]] = []
    cost_rows: list[dict[str, str | float]] = []
    today = date.today().isoformat()

    if vercel_token:
        status, teams_payload = _http_json(
            "https://api.vercel.com/v2/teams?limit=20",
            headers={"Authorization": f"Bearer {vercel_token}"},
        )
        team = None
        if isinstance(teams_payload, dict):
            teams = teams_payload.get("teams", [])
            for item in teams:
                if item.get("slug") == "dc-joinery":
                    team = item
                    break
        if team:
            plan = (((team.get("billing") or {}).get("plan")) or "unknown").lower()
            detail = f"Team {team.get('slug')} plan={plan}"
            audit_rows.append(("Vercel", "ok", detail, checked_at))
            if plan == "hobby":
                cost_rows.append(
                    {
                        "date": today,
                        "vendor": "Vercel",
                        "service": "Hosting plan (Hobby)",
                        "category": "Hosting",
                        "billing_cycle": "monthly",
                        "unit_cost_gbp": 0,
                        "qty": 1,
                        "notes": f"Plan detected from Vercel API ({period_days}-day audit window).",
                        "verification_status": "verified",
                        "evidence": "Vercel team billing.plan=hobby",
                    }
                )
            else:
                cost_rows.append(
                    {
                        "date": today,
                        "vendor": "Vercel",
                        "service": "Hosting plan",
                        "category": "Hosting",
                        "billing_cycle": "monthly",
                        "unit_cost_gbp": 0,
                        "qty": 1,
                        "notes": f"Plan detected ({period_days}-day audit window), amount must come from invoice.",
                        "verification_status": "pending",
                        "evidence": detail,
                    }
                )
        else:
            audit_rows.append(("Vercel", "error", f"status={status} payload={teams_payload}", checked_at))
    else:
        audit_rows.append(("Vercel", "pending", "Missing VERCEL_TOKEN for billing audit.", checked_at))

    providers = [
        ("Supabase", "SUPABASE_URL", "Hosting", "Database + storage"),
        ("Resend", "RESEND_API_KEY", "Email", "Transactional email"),
        ("OpenAI", "OPENAI_API_KEY", "AI/API", "Model usage"),
        ("Replicate", "REPLICATE_API_TOKEN", "AI/API", "Image generation usage"),
        ("PDFShift", "PDFSHIFT_API_KEY", "Tools", "PDF generation usage"),
        ("Domain Registrar", "DOMAIN_INVOICE_REFERENCE", "Domain", "dcjoineryni.uk domain"),
    ]
    for provider, env_key, category, service in providers:
        if env.get(env_key):
            status_label = "pending"
            detail = "Credential present, invoice/usage amount still required."
            if provider == "OpenAI":
                now_ts = int(datetime.utcnow().timestamp())
                since_ts = now_ts - period_days * 24 * 3600
                query = urllib.parse.urlencode({"start_time": since_ts, "end_time": now_ts})
                status, payload = _http_json(
                    f"https://api.openai.com/v1/organization/costs?{query}",
                    headers={"Authorization": f"Bearer {env['OPENAI_API_KEY']}"},
                )
                if status == 200:
                    status_label = "ok"
                    detail = f"Usage endpoint accessible for last {period_days} days."
                else:
                    detail = f"status={status} details={payload}"
            elif provider == "Resend":
                status_label, detail = _probe_resend_usage(env["RESEND_API_KEY"], period_start)
            elif provider == "Replicate":
                status_label, detail = _probe_replicate_usage(env["REPLICATE_API_TOKEN"], period_start)
            elif provider == "Supabase":
                supabase_key = env.get("SUPABASE_SERVICE_ROLE_KEY", "")
                if supabase_key:
                    status_label, detail = _probe_supabase_usage(env["SUPABASE_URL"], supabase_key, period_start)
            elif provider == "PDFShift":
                status, payload, auth_mode = _pdfshift_audit(env["PDFSHIFT_API_KEY"])
                if status == 200:
                    status_label = "ok"
                detail = f"status={status} mode={auth_mode} details={payload}"
            audit_rows.append((provider, status_label, detail, checked_at))
            cost_rows.append(
                {
                    "date": today,
                    "vendor": provider,
                    "service": service,
                    "category": category,
                    "billing_cycle": "monthly" if category != "Domain" else "yearly",
                    "unit_cost_gbp": 0,
                    "qty": 1,
                    "notes": "Real amount not inferred. Fill from invoice/export.",
                    "verification_status": "pending",
                    "evidence": detail,
                }
            )
        else:
            audit_rows.append((provider, "missing", f"Missing env key: {env_key}", checked_at))
            cost_rows.append(
                {
                    "date": today,
                    "vendor": provider,
                    "service": service,
                    "category": category,
                    "billing_cycle": "monthly" if category != "Domain" else "yearly",
                    "unit_cost_gbp": 0,
                    "qty": 1,
                    "notes": "Missing credential/invoice source.",
                    "verification_status": "missing",
                    "evidence": f"Missing env key: {env_key}",
                }
            )

    audit_rows, cost_rows = _merge_verified_costs(previous_payload, audit_rows, cost_rows)

    payload = {
        "checked_at": checked_at,
        "period_days": period_days,
        "audit_rows": [
            {"provider": p, "status": s, "details": d, "checked_at": t}
            for p, s, d, t in audit_rows
        ],
        "cost_rows": cost_rows,
    }
    AUX_AUDIT_PATH.write_text(json.dumps(payload, indent=2))
    print(f"Audit snapshot written: {AUX_AUDIT_PATH}")


def cmd_add(args: argparse.Namespace) -> None:
    ensure_workbook(WORKBOOK_PATH)
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Costs"]
    append_cost_row(
        ws=ws,
        entry_date=parse_date(args.date),
        vendor=args.vendor,
        service=args.service,
        category=args.category,
        cycle=args.billing_cycle,
        unit_cost=float(args.unit_cost_gbp),
        qty=float(args.qty),
        notes=args.notes or "",
        verification_status=args.verification_status,
        evidence=args.evidence,
    )
    wb.save(WORKBOOK_PATH)
    print(f"Cost added to {WORKBOOK_PATH}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="DC Joinery cost tracker workbook helper")
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_parser = subparsers.add_parser("init", help="Create fresh workbook with dashboard")
    init_parser.set_defaults(func=lambda _: cmd_init())

    audit_parser = subparsers.add_parser("audit", help="Create provider audit snapshot (JSON)")
    audit_parser.add_argument("--vercel-token", help="Vercel API token (or use VERCEL_TOKEN env)")
    audit_parser.add_argument("--days", type=int, default=28, help="Rolling audit window in days (default: 28)")
    audit_parser.set_defaults(func=cmd_audit)

    add_parser = subparsers.add_parser("add", help="Append one cost entry")
    add_parser.add_argument("--date", help="Entry date YYYY-MM-DD (default: today)")
    add_parser.add_argument("--vendor", required=True)
    add_parser.add_argument("--service", required=True)
    add_parser.add_argument("--category", required=True, choices=CATEGORIES)
    add_parser.add_argument("--billing-cycle", required=True, choices=["monthly", "yearly", "one-time"])
    add_parser.add_argument("--unit-cost-gbp", required=True, type=float)
    add_parser.add_argument("--qty", default=1, type=float)
    add_parser.add_argument("--notes", default="")
    add_parser.add_argument("--verification-status", default="manual", choices=["verified", "pending", "missing", "manual"])
    add_parser.add_argument("--evidence", default="")
    add_parser.set_defaults(func=cmd_add)

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
